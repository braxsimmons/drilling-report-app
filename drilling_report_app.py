
import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import json
from datetime import datetime, date, time
from typing import Any, Dict, List, Optional, Tuple
import zipfile

def to_float(x: Any) -> Optional[float]:
    if x is None or (isinstance(x, str) and x.strip() == ""):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace(",", " ").strip()
    m = re.findall(r"[-+]?\d*\.\d+|[-+]?\d+", s)
    return float(m[0]) if m else None

def to_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()

def ensure_cols(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for c in cols:
        if c not in df.columns:
            df[c] = None
    return df[cols]

def default_df(columns: List[str], rows: int = 1) -> pd.DataFrame:
    return pd.DataFrame([{c: None for c in columns} for _ in range(rows)])

st.set_page_config(page_title="Daily Drilling Report — Data Entry & Normalizer", layout="wide")
st.title("🛢️ Daily Drilling Report — Data Entry & Normalizer")

with st.expander("Template (optional) → Upload the Excel you want to fill.", expanded=False):
    template_file = st.file_uploader("Excel template (.xlsx/.xlsm). If provided, the app will attempt to write values into matching labels and table headers.", type=["xlsx","xlsm"])

st.markdown("---")
st.subheader("Header")
c1, c2, c3, c4 = st.columns(4)
report_number = c1.text_input("Report #", value="1")
report_date   = c2.date_input("Date", value=datetime.today().date())
oil_company   = c3.text_input("Oil Company", value="KIMMERIDGE TEXAS GAS")
contractor    = c4.text_input("Contractor", value="Primo")

c5, c6, c7, c8 = st.columns(4)
well_name     = c5.text_input("Well Name & No", value="HUFF A 106H")
location      = c6.text_input("Location", value="MCMULLEN")
county_state  = c7.text_input("County, State", value="TEXAS")
permit_no     = c8.text_input("Permit Number", value="909899")

c9, c10 = st.columns(2)
api_number    = c9.text_input("API #", value="42-311-37631")
rig_con_no    = c10.text_input("Rig Contractor & No.", value="7")

st.markdown("---")
st.subheader("Daily Progress")
c1, c2, c3, c4 = st.columns(4)
depth_0000 = c1.number_input("00:00 Depth (ft)", value=120.0, step=1.0, format="%.2f")
depth_2400 = c2.number_input("24:00 Depth (ft)", value=288.0, step=1.0, format="%.2f")
footage_today = c3.number_input("Footage Today (ft)", value=168.0, step=1.0, format="%.2f")
rop_today     = c4.number_input("ROP Today (ft/hr)", value=56.0, step=0.1, format="%.2f")

c5, c6, c7, c8 = st.columns(4)
drlg_hrs_today = c5.number_input("Drlg Hrs Today", value=3.0, step=0.25, format="%.2f")
current_run_ftg= c6.number_input("Current Run Ftg", value=0.0, step=1.0, format="%.2f")
circ_hrs_today = c7.number_input("Circ Hrs Today", value=0.0, step=0.25, format="%.2f")

st.markdown("---")
colA, colB = st.columns(2, gap="large")

with colA:
    st.subheader("Pump Data")
    pump_cols = ["pump_no","bbls_per_stroke","gals_per_stroke","vol_gpm","spm"]
    pump_df = st.data_editor(
        default_df(pump_cols, rows=2).assign(pump_no=[1,2]),
        num_rows="dynamic",
        use_container_width=True,
        key="pump"
    )

    st.subheader("Mud Data")
    mud_cols = ["weight_ppg","viscosity_sec","pressure_psi"]
    mud_df = st.data_editor(default_df(mud_cols), use_container_width=True, key="mud")

    st.subheader("Drilling Parameters & Motor")
    param_cols = [
        "st_wt_rot_klbs","pu_wt_klbs","so_wt_klbs","wob_klbs",
        "rotary_rpm","motor_rpm","total_bit_rpm",
        "rot_tq_off_btm_ftlb","rot_tq_on_btm_ftlb",
        "off_bottom_pressure_psi","on_bottom_pressure_psi"
    ]
    params_df = st.data_editor(default_df(param_cols), use_container_width=True, key="params")

    motor_cols = [
        "run_no","size_in","type","serial_no","tool_deflection","avg_diff_press_psi",
        "daily_drill_hrs","daily_circ_hrs","daily_total_hrs","acc_drill_hrs","acc_circ_hrs",
        "depth_in_ft","depth_out_ft"
    ]
    motor_df = st.data_editor(default_df(motor_cols), use_container_width=True, key="motor")

with colB:
    st.subheader("BHA")
    bha_cols = ["item","od_in","id_in","weight","connection","length_ft","depth_ft"]
    bha_df = st.data_editor(
        default_df(bha_cols, rows=8).assign(item=["BIT","MOTOR","UBHO","MONEL","SHOCK SUB","8\" DC","XO","6\" DC"]),
        num_rows="dynamic",
        use_container_width=True,
        key="bha"
    )

    st.subheader("Survey Info (optional)")
    survey_cols = ["depth_ft","inc_deg","azi_deg"]
    survey_df = st.data_editor(default_df(survey_cols, rows=3), num_rows="dynamic", use_container_width=True, key="survey")

st.markdown("---")
st.subheader("Bit Data")
bit_cols = ["no","size_in","mfg","type","nozzles_or_tfa","serial_no","depth_in_ft","cum_footage_ft","cum_hours","depth_out_ft","dull_ir","dull_or","dull_dc","loc","bs","g_16","oc","rpld"]
bit_df = st.data_editor(default_df(bit_cols, rows=2), num_rows="dynamic", use_container_width=True, key="bit")

st.markdown("---")
colC, colD, colE = st.columns(3, gap="large")

with colC:
    st.subheader("Casing")
    casing_cols = ["od_in","id_in","weight","grade","connection","depth_set_ft"]
    casing_df = st.data_editor(default_df(casing_cols, rows=3), num_rows="dynamic", use_container_width=True, key="casing")

with colD:
    st.subheader("Drill Pipe")
    dp_cols = ["od_in","id_in","weight","grade","connection"]
    drillpipe_df = st.data_editor(default_df(dp_cols, rows=3), num_rows="dynamic", use_container_width=True, key="drillpipe")

with colE:
    st.subheader("Rental Equipment")
    rental_cols = ["item","serial_no","date_received","date_returned"]
    rental_df = st.data_editor(default_df(rental_cols, rows=2), num_rows="dynamic", use_container_width=True, key="rental")

st.markdown("---")
st.subheader("Time Breakdown & Forecast")
tb_cols = ["from_time","to_time","hrs","start_depth_ft","end_depth_ft","cl","description","code","forecast"]
time_df = st.data_editor(default_df(tb_cols, rows=6), num_rows="dynamic", use_container_width=True, key="timebk")

st.markdown("---")
colF, colG = st.columns(2, gap="large")

with colF:
    st.subheader("Fuel")
    fuel_cols = ["fuel_type","vendor","begin_qty","received","total","used","remaining"]
    fuel_df = st.data_editor(default_df(fuel_cols, rows=1).assign(fuel_type="DIESEL"), use_container_width=True, key="fuel")

    st.subheader("Chemicals (optional)")
    chem_cols = ["additive","qty","unit"]
    chemicals_df = st.data_editor(default_df(chem_cols, rows=5), num_rows="dynamic", use_container_width=True, key="chems")

with colG:
    st.subheader("Personnel / Crew")
    personnel_cols = ["tour","role","name","hours"]
    personnel_df = st.data_editor(default_df(personnel_cols, rows=8), num_rows="dynamic", use_container_width=True, key="people")

st.markdown("---")
if st.button("Normalize & Export"):
    core = dict(
        report_number=to_str(report_number),
        date=str(report_date),
        oil_company=to_str(oil_company),
        contractor=to_str(contractor),
        well_name=to_str(well_name),
        location=to_str(location),
        county_state=to_str(county_state),
        permit_number=to_str(permit_no),
        api_number=to_str(api_number),
        rig_contractor_no=to_str(rig_con_no),
        depth_0000_ft=to_float(depth_0000),
        depth_2400_ft=to_float(depth_2400),
        footage_today_ft=to_float(footage_today),
        rop_today_ft_hr=to_float(rop_today),
        drlg_hrs_today=to_float(drlg_hrs_today),
        current_run_ftg=to_float(current_run_ftg),
        circ_hrs_today=to_float(circ_hrs_today),
        created_at=str(datetime.utcnow())
    )

    def add_meta(df: pd.DataFrame) -> pd.DataFrame:
        meta = pd.DataFrame([core] * len(df.index))
        return pd.concat([meta.reset_index(drop=True), df.reset_index(drop=True)], axis=1)

    def normalize_numeric(df: pd.DataFrame):
        for col in df.columns:
            if any(k in col.lower() for k in ["ft","hrs","rpm","psi","qty","no","weight","od","id","gpm","spm","size","length","depth","bbl","cum","g_16"]):
                df[col] = df[col].map(to_float)
            else:
                df[col] = df[col].map(to_str)
        return df

    pump_df[:] = normalize_numeric(pump_df)
    mud_df[:] = normalize_numeric(mud_df)
    params_df[:] = normalize_numeric(params_df)
    motor_df[:] = normalize_numeric(motor_df)
    bha_df[:] = normalize_numeric(bha_df)
    survey_df[:] = normalize_numeric(survey_df)
    bit_df[:] = normalize_numeric(bit_df)
    casing_df[:] = normalize_numeric(casing_df)
    drillpipe_df[:] = normalize_numeric(drillpipe_df)
    rental_df[:] = normalize_numeric(rental_df)
    time_df[:] = normalize_numeric(time_df)
    fuel_df[:] = normalize_numeric(fuel_df)
    chemicals_df[:] = normalize_numeric(chemicals_df)
    personnel_df[:] = normalize_numeric(personnel_df)

    outputs = {
        "core.csv": pd.DataFrame([core]),
        "pump.csv": add_meta(ensure_cols(pump_df, ["pump_no","bbls_per_stroke","gals_per_stroke","vol_gpm","spm"])),
        "mud.csv": add_meta(ensure_cols(mud_df, ["weight_ppg","viscosity_sec","pressure_psi"])),
        "params.csv": add_meta(ensure_cols(params_df, ["st_wt_rot_klbs","pu_wt_klbs","so_wt_klbs","wob_klbs","rotary_rpm","motor_rpm","total_bit_rpm","rot_tq_off_btm_ftlb","rot_tq_on_btm_ftlb","off_bottom_pressure_psi","on_bottom_pressure_psi"])),
        "motor.csv": add_meta(ensure_cols(motor_df, ["run_no","size_in","type","serial_no","tool_deflection","avg_diff_press_psi","daily_drill_hrs","daily_circ_hrs","daily_total_hrs","acc_drill_hrs","acc_circ_hrs","depth_in_ft","depth_out_ft"])),
        "bha.csv": add_meta(ensure_cols(bha_df, ["item","od_in","id_in","weight","connection","length_ft","depth_ft"])),
        "survey.csv": add_meta(ensure_cols(survey_df, ["depth_ft","inc_deg","azi_deg"])),
        "bit.csv": add_meta(ensure_cols(bit_df, ["no","size_in","mfg","type","nozzles_or_tfa","serial_no","depth_in_ft","cum_footage_ft","cum_hours","depth_out_ft","dull_ir","dull_or","dull_dc","loc","bs","g_16","oc","rpld"])),
        "casing.csv": add_meta(ensure_cols(casing_df, ["od_in","id_in","weight","grade","connection","depth_set_ft"])),
        "drillpipe.csv": add_meta(ensure_cols(drillpipe_df, ["od_in","id_in","weight","grade","connection"])),
        "rental_equipment.csv": add_meta(ensure_cols(rental_df, ["item","serial_no","date_received","date_returned"])),
        "time_breakdown.csv": add_meta(ensure_cols(time_df, ["from_time","to_time","hrs","start_depth_ft","end_depth_ft","cl","description","code","forecast"])),
        "fuel.csv": add_meta(ensure_cols(fuel_df, ["fuel_type","vendor","begin_qty","received","total","used","remaining"])),
        "chemicals.csv": add_meta(ensure_cols(chemicals_df, ["additive","qty","unit"])),
        "personnel.csv": add_meta(ensure_cols(personnel_df, ["tour","role","name","hours"])),
    }

    payload = {"core": core}
    for name, dfv in outputs.items():
        if name == "core.csv":
            continue
        payload[name.replace(".csv","")] = json.loads(dfv.to_json(orient="records"))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w") as zf:
        for fname, dfv in outputs.items():
            zf.writestr(fname, dfv.to_csv(index=False))
        zf.writestr("normalized.json", json.dumps(payload, indent=2))

        ddl = """
        -- Example DDL for Supabase/Postgres. Adjust types and PKs as needed.
        CREATE TABLE IF NOT EXISTS drilling_core (
            report_number text,
            date date,
            oil_company text,
            contractor text,
            well_name text,
            location text,
            county_state text,
            permit_number text,
            api_number text,
            rig_contractor_no text,
            depth_0000_ft numeric,
            depth_2400_ft numeric,
            footage_today_ft numeric,
            rop_today_ft_hr numeric,
            drlg_hrs_today numeric,
            current_run_ftg numeric,
            circ_hrs_today numeric,
            created_at timestamptz
        );
        """
        zf.writestr("schema_example.sql", ddl)

        # Try filling template minimally (labels only) if provided
        if template_file is not None:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(template_file, keep_vba=True)
                ws = wb.active
                def write_next_to(label: str, value: Any):
                    label_norm = re.sub(r"\s+"," ",label.strip().lower())
                    for row in ws.iter_rows(values_only=False):
                        for cell in row:
                            val = cell.value
                            if isinstance(val, str) and label_norm in re.sub(r"\s+"," ",val.strip().lower()):
                                c = cell.column + 1
                                tgt = ws.cell(row=cell.row, column=c)
                                tgt.value = value
                                return True
                    return False
                write_next_to("Report #", core["report_number"])
                write_next_to("Date", core["date"])
                write_next_to("Oil Company", core["oil_company"])
                write_next_to("Contractor", core["contractor"])
                write_next_to("Well Name & No", core["well_name"])
                write_next_to("Location", core["location"])
                write_next_to("County, State", core["county_state"])
                write_next_to("Permit Number", core["permit_number"])
                write_next_to("API #", core["api_number"])
                write_next_to("Rig Contractor & No.", core["rig_contractor_no"])
                write_next_to("00:00 Depth", core["depth_0000_ft"])
                write_next_to("24:00 Depth", core["depth_2400_ft"])
                write_next_to("Footage Today", core["footage_today_ft"])
                write_next_to("Drlg Hrs Today", core["drlg_hrs_today"])
                write_next_to("ROP Today", core["rop_today_ft_hr"])
                write_next_to("Current Run Ftg", core["current_run_ftg"])
                write_next_to("Circ Hrs Today", core["circ_hrs_today"])

                out_xlsx = io.BytesIO()
                wb.save(out_xlsx)
                zf.writestr("filled_template.xlsx", out_xlsx.getvalue())
            except Exception as e:
                zf.writestr("template_error.txt", f"Failed to fill template: {e}")

    zip_buf.seek(0)
    st.download_button("⬇️ Download All Outputs (ZIP)", data=zip_buf.getvalue(), file_name=f"drilling_report_outputs_{report_date}.zip", mime="application/zip")
    st.success("Done! ZIP contains separate CSVs, normalized.json, schema_example.sql, and filled_template.xlsx (if provided).")
